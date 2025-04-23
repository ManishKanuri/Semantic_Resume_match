import json
import requests
from openpyxl import load_workbook
from pathlib import Path

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL = "llama3"

def load_resume_text(resume_path):
    text = Path(resume_path).read_text()
    json_part = text[text.index("{"):text.rindex("}")+1]
    resume = json.loads(json_part)
    summary = resume.get("summary", "")
    skills = ", ".join(resume.get("skills", []))
    experience = "\n".join(
        f"{e.get('title', '')} at {e.get('company', '')}: {'; '.join(e.get('details', []))}"
        for e in resume.get("experience", [])
    )
    return f"{summary}\nSkills: {skills}\nExperience:\n{experience}"

def load_jd_text(jd_path):
    return Path(jd_path).read_text()

def get_explanation(jd_text, resume_text):
    prompt = f"""
You are an expert recruiter. Given the following job description and resume, explain in 2-3 bullet points why this resume is a good match for the job.

Job Description:
{jd_text}

Resume:
{resume_text}

Explain why this candidate is a good match:
"""
    response = requests.post(OLLAMA_URL, json={
        "model": MODEL,
        "prompt": prompt,
        "stream": False
    })
    return response.json()['response'].strip()

def add_explanations_to_excel(file_path="jd_resume_match_results.xlsx", resume_base="resume_generator/resume_data", jd_base="jd_data/raw_jd_texts"):
    wb = load_workbook(file_path)
    ws = wb.active

    # Add "Explanation" column header if not present
    if "Explanation" not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=ws.max_column + 1).value = "Explanation"
        explanation_col = ws.max_column
    else:
        explanation_col = [cell.value for cell in ws[1]].index("Explanation") + 1

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        jd_name = row[0].value
        resume_file = row[3].value
        jd_path = Path(jd_base) / f"{jd_name}.txt"
        resume_path = Path(resume_base) / resume_file

        try:
            jd_text = load_jd_text(jd_path)
            resume_text = load_resume_text(resume_path)
            explanation = get_explanation(jd_text, resume_text)
        except Exception as e:
            explanation = f"⚠️ Error: {e}"

        # ✅ Write explanation into correct cell using row/col references
        ws.cell(row=row[0].row, column=explanation_col).value = explanation
        print(f"✅ Processed: {resume_file} for {jd_name}")

    wb.save(file_path)
    print("✅ Explanations added and file saved!")

# ▶️ Run
if __name__ == "__main__":
    add_explanations_to_excel()


export ANTHROPIC_API_KEY="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"
import anthropic

anthropic_client = anthropic.Anthropic()

def get_claude_explanation(jd_text, resume_text):
    prompt = f"""
You are an expert recruiter. Given the following job description and resume, explain in 2-3 bullet points why this resume is a good match for the job.

Job Description:
{jd_text}

Resume:
{resume_text}

Explain why this candidate is a good match:
"""
    response = anthropic_client.messages.create(
        model="claude-3-sonnet-20240229",  # or claude-3-opus-20240229
        max_tokens=500,
        temperature=0.7,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )
    return response.content[0].text.strip()
