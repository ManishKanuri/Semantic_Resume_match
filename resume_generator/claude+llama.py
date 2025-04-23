import json
import requests
from openpyxl import load_workbook
from pathlib import Path

OLLAMA_URL = "http://localhost:11434/api/generate"
LLAMA_MODEL = "llama3"
MISTRAL_MODEL = "mistral:instruct"

EXCEL_PATH = "jd_resume_match_results_final.xlsx"
RESUME_DIR = Path("resume_generator/resume_data")
JD_DIR = Path("jd_data/raw_jd_texts")


def load_resume_text(resume_path):
    text = Path(resume_path).read_text()
    json_part = text[text.index("{"):text.rindex("}") + 1]
    resume = json.loads(json_part)
    summary = resume.get("summary", "")
    skills = ", ".join(resume.get("skills", []))
    experience = "\n".join(
        f"{e.get('title', '')} at {e.get('company', '')}: {'; '.join(e.get('details', []))}"
        for e in resume.get("experience", [])
    )
    return f"{summary}\nSkills: {skills}\nExperience:\n{experience}"


def load_jd_text(jd_path):
    return Path(jd_path).read_text()


def get_explanation(model_name, jd_text, resume_text):
    prompt = f"""
You are an expert recruiter. Given the following job description and resume, explain in 2-3 bullet points why this resume is a good match for the job.

Job Description:
{jd_text}

Resume:
{resume_text}

Explain why this candidate is a good match:
"""
    try:
        response = requests.post(OLLAMA_URL, json={
            "model": model_name,
            "prompt": prompt,
            "stream": False
        })
        return response.json()['response'].strip()
    except Exception as e:
        return f"Error: {e}"


def add_explanations_to_excel(file_path=EXCEL_PATH, resume_base=RESUME_DIR, jd_base=JD_DIR):
    wb = load_workbook(file_path)
    ws = wb.active

    # Ensure both columns exist
    headers = [cell.value for cell in ws[1]]
    if "LLaMA Explanation" not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = "LLaMA Explanation"
        llama_col = ws.max_column
        ws.cell(row=1, column=llama_col + 1).value = "Mistral Explanation"
        mistral_col = llama_col + 1
    else:
        llama_col = headers.index("LLaMA Explanation") + 1
        mistral_col = headers.index("Mistral Explanation") + 1

    for row in ws.iter_rows(min_row=2):
        jd_name = row[0].value
        resume_file = row[3].value
        jd_path = Path(jd_base) / f"{jd_name}.txt"
        resume_path = Path(resume_base) / resume_file

        try:
            jd_text = load_jd_text(jd_path)
            resume_text = load_resume_text(resume_path)

            llama_expl = get_explanation(LLAMA_MODEL, jd_text, resume_text)
            mistral_expl = get_explanation(MISTRAL_MODEL, jd_text, resume_text)

            # Write to cells using index math (1-based column indexing)
            ws.cell(row=row[0].row, column=llama_col).value = llama_expl
            ws.cell(row=row[0].row, column=mistral_col).value = mistral_expl

            print(f"{resume_file} for JD {jd_name} — explanations added")

        except Exception as e:
            print(f"Error processing {resume_file} — {e}")

    wb.save(file_path)
    print(f"\nSaved explanations to {file_path}")


# ▶️ Run
if __name__ == "__main__":
    add_explanations_to_excel()
