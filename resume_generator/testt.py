import json
import re
import time
import faiss
import torch
import numpy as np
from pathlib import Path
from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook, load_workbook

OUTPUT_FILE = "jd_resume_match_top_results.xlsx"

def extract_json_block(text):
    try:
        return text[text.index("{"):text.rindex("}")+1]
    except:
        return None

def estimate_experience_years(experience):
    total_years = 0
    for job in experience:
        duration = job.get("duration", "")
        match = re.search(r"(\d{4})\s*[-–—to]+\s*(\d{4}|present)", duration, re.I)
        if match:
            start = int(match.group(1))
            end = int(match.group(2)) if match.group(2).isdigit() else 2025
            total_years += max(0, end - start)
        elif "summer" in duration.lower() and re.search(r"\d{4}", duration):
            total_years += 0.25
    return total_years

def extract_text_from_resume(resume):
    summary = resume.get("summary", "")
    skills = " ".join(resume.get("skills", []))
    exp_text = " ".join(
        f"{e.get('title', '')} at {e.get('company', '')}. {' '.join(e.get('details', []))}"
        for e in resume.get("experience", [])
    )
    education = f"{resume.get('education', {}).get('degree', '')} from {resume.get('education', {}).get('institution', '')}"
    certs = " ".join(resume.get("certifications", []))
    return f"{summary} {skills} {exp_text} {education} {certs}"

def append_to_excel(rows, clear=False):
    path = Path(OUTPUT_FILE)
    if path.exists() and not clear:
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["JD Name", "Method", "Rank", "Resume", "Score/Distance", "Time Taken (s)"])

    for row in rows:
        ws.append(row)

    wb.save(OUTPUT_FILE)

def match_resumes_to_jd(jd_txt_path, clear_previous=False):
    model = SentenceTransformer("all-MiniLM-L6-v2")
    resume_dir = Path(__file__).parent / "resume_data"

    jd_path = Path(jd_txt_path)
    jd_name = jd_path.stem
    jd_text = jd_path.read_text()

    parsed_path = Path("jd_data/parsed") / f"{jd_name}.json"
    min_exp, max_exp = None, None  # Default to no experience requirement
    if parsed_path.exists():
        jd_struct = json.load(open(parsed_path))
        min_exp = jd_struct.get("min_experience", None)
        max_exp = jd_struct.get("max_experience", None)
        print(f"Required: Min Exp: {min_exp}, Max Exp: {max_exp}")
    else:
        print(f"⚠️ No parsed JD metadata found at {parsed_path}")

    jd_embed = model.encode(jd_text, convert_to_tensor=True)
    jd_embed_np = jd_embed.cpu().numpy().reshape(1, -1).astype(np.float32)

    embeddings, metadata = [], []
    for path in resume_dir.glob("*.json"):
        try:
            content = path.read_text()
            json_text = extract_json_block(content)
            resume = json.loads(json_text)
            exp_years = estimate_experience_years(resume.get("experience", []))
            
            # Only filter if min_exp or max_exp is not None
            if (min_exp is not None and exp_years < min_exp) or (max_exp is not None and exp_years > max_exp):
                print(f"⛔ Skipping {path.name} — {exp_years} years (out of bounds)")
                continue
                
            text = extract_text_from_resume(resume)
            emb = model.encode(text, normalize_embeddings=False)
            embeddings.append(emb)
            metadata.append(path.name)
        except Exception as e:
            print(f"⚠️ Skipping {path.name} — {e}")

    if not embeddings:
        print("❌ No resumes matched the experience criteria.")
        return

    embeddings_np = np.vstack(embeddings).astype(np.float32)

    # Vanilla Cosine
    start1 = time.time()
    cosine_scores = [
        (metadata[i], float(util.cos_sim(jd_embed, torch.tensor(embeddings_np[i:i+1]))[0]))
        for i in range(len(embeddings_np))
    ]
    cosine_scores.sort(key=lambda x: x[1], reverse=True)
    end1 = time.time()

    # FAISS Cosine
    start2 = time.time()
    norm_embeds = np.require(embeddings_np, dtype=np.float32, 
                        requirements=['C_CONTIGUOUS', 'ALIGNED', 'WRITEABLE'])
    faiss.normalize_L2(norm_embeds)

    jd_norm = np.require(jd_embed_np, dtype=np.float32, 
                    requirements=['C_CONTIGUOUS', 'ALIGNED', 'WRITEABLE'])
    faiss.normalize_L2(jd_norm)

    faiss_cos_index = faiss.IndexFlatIP(norm_embeds.shape[1])
    faiss_cos_index.add(norm_embeds)
    D_cos, I_cos = faiss_cos_index.search(jd_norm, min(5, len(norm_embeds)))
    faiss_cosine_results = [(metadata[idx], float(D_cos[0][i])) for i, idx in enumerate(I_cos[0])]
    end2 = time.time()

    # FAISS L2
    start3 = time.time()
    faiss_l2_index = faiss.IndexFlatL2(embeddings_np.shape[1])
    faiss_l2_index.add(embeddings_np)
    D_l2, I_l2 = faiss_l2_index.search(jd_embed_np, min(5, len(embeddings_np)))
    faiss_l2_out = [(metadata[idx], float(D_l2[0][i])) for i, idx in enumerate(I_l2[0])]
    end3 = time.time()

    # Write rows with safety for fewer than 5 results
    rows = []
    max_results = min(5, len(cosine_scores), len(faiss_cosine_results), len(faiss_l2_out))
    
    for rank in range(max_results):
        rows.append([
            jd_name, 
            "Vanilla Cosine", 
            rank+1, 
            cosine_scores[rank][0],
            cosine_scores[rank][1], 
            round(end1 - start1, 4)
        ])
        rows.append([
            jd_name, 
            "FAISS Cosine", 
            rank+1, 
            faiss_cosine_results[rank][0],
            faiss_cosine_results[rank][1],
            round(end2 - start2, 4)
        ])
        rows.append([
            jd_name, 
            "FAISS L2", 
            rank+1, 
            faiss_l2_out[rank][0],
            faiss_l2_out[rank][1], 
            round(end3 - start3, 4)
        ])

    append_to_excel(rows, clear=clear_previous)
    print(f"\n✅ Results for {jd_name} written to Excel.")
    
    # Print top matches summary
    print(f"Found {len(cosine_scores)} matching resumes. Top matches:")
    for i, (resume, score) in enumerate(cosine_scores[:3], 1):
        print(f"{i}. {resume}: {score:.4f}")

# ▶️ Run
if __name__ == "__main__":
    # Set clear_previous=True to overwrite previous results (prevents duplicates)
    match_resumes_to_jd("jd_data/raw_jd_texts/computer_vision_visualization.txt", clear_previous=True)
